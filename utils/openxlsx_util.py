#! /usr/bin/env python
# -*- conding:utf-8 -*-
"""
@Time : 2020/5/3 21:34
@Author : 搁浅灬惜缘
@file : excel_util.py
"""

import openpyxl
import os
from utils.logging_util import Logger

logging = Logger().get_logger()


class OpenxlsxUtil(object):
    def __init__(self, open_file=None, index=0):
        '''
        需传入文件名称和sheet表格，默认为0
        '''
        self.open_file = open_file
        self.index = index

    def get_file_path(self):
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if self.open_file is None:
            file_path = os.path.join(base_dir + '/data/' + 'template.xlsx')
        else:
            file_path = os.path.join(base_dir + '/data/' + self.open_file)
        return file_path

    def get_openxlsx(self):
        open_excel = None
        file_path = self.get_file_path()
        try:
            logging.info("获取Excel-xlsx文件:{0}".format(file_path))
            open_excel = openpyxl.load_workbook(file_path)
        except:
            logging.error("获取Excel文件失败")
        Logger().close_logger()
        # open_excel.close()
        return open_excel

    def get_sheet(self):
        '''
        获取文件完整路径
        :return:
        '''
        table = None
        open_excel = self.get_openxlsx()
        try:
            sheet_names = open_excel.sheetnames
            logging.info("获取Excel-xlsx的第{0}个sheet".format(self.index + 1))
            table = open_excel[sheet_names[self.index]]
        except:
            logging.error("获取Excel-xlsx的第{0}个sheet失败".format(self.index + 1))
        Logger().close_logger()
        # open_excel.close()
        return table

    def get_rows(self):
        '''
        获取Excel文件总行数
        :return:
        '''
        row = None
        table = self.get_sheet()
        try:
            logging.info("获取Excel总行数:{0}".format(table.max_row))
            row = table.max_row
        except:
            logging.error("获取Excel总行数失败")
        Logger().close_logger()
        return row

    def get_row_datas(self, row):
        '''
        获取文件数据内容
        :return:
        '''
        result = []
        table = self.get_sheet()
        try:
            for i in table[row]:
                result.append(i.value)
            logging.info("获取Excel-xlsx第{0}行数据为{1}".format(row, result))
        except:
            logging.error("获取Excel文件内容失败")
        Logger().close_logger()
        return result

    def get_cols(self):
        col = None
        table = self.get_sheet()
        try:
            logging.info("获取Excel总列数:{0}".format(table.max_column))
            col = table.max_column
        except:
            logging.error("获取Excel总列数失败")
        Logger().close_logger()
        return col

    def get_col_datas(self, col):
        result = []
        table = self.get_sheet()
        row = self.get_rows()
        try:
            for i in range(1, row + 1):
                result.append(table.cell(i, col).value)
            logging.info("获取Excel-xlsx第{0}列数据为{1}".format(col, result))
        except:
            logging.error("获取Excel文件内容失败")
        Logger().close_logger()
        return result

    def get_data_value(self, row, col):
        '''
        获取单元格的值
        :param row: 行号
        :param col: 列号
        :return: 单元格值
        '''
        data = None
        table = self.get_sheet()
        try:
            logging.info("获取Excel第{0}行第{1}列的值:{2}".format(row, col, table.cell(row, col).value))
            data = table.cell(row, col).value
        except:
            logging.error("获取Excel单元格值失败")
        Logger().close_logger()
        return data

    def write_value(self, row, col, value):
        open_excel = self.get_openxlsx()
        try:
            logging.info("获取Excel-xlsx的第{0}个sheet表".format(self.index + 1))
            sheet = open_excel.worksheets[self.index]
            logging.info("在工作表{0}中第{1}行第{2}列中写入值:{3}".format(sheet.title, row, col, value))
            # logging.info("在工作表{0}中第{1}行第{2}列中写入值:{3}".format(open_excel.sheetnames[self.index], row, col, value))
            sheet.cell(row, col).value = value
            file_path = self.get_file_path()
            logging.info("保存文件:{0}".format(file_path))
            open_excel.save(file_path)
            open_excel.close()
        except:
            logging.error("Excel中写入数据失败")
        Logger().close_logger()


if __name__ == '__main__':
    x = OpenxlsxUtil()
    # t = x.get_openxlsx()
    print(x.write_value(3, 6, 'qwert'))
    # open_excel.save(open_path)
